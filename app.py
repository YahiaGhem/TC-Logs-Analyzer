import io
import re
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Complaint Workbook Generator", layout="wide")

st.title("Complaint Workbook Generator")
st.write(
    "Upload the previously generated output sheet (optional — skip on first run), "
    "then upload one or more new raw complaint Excel files. "
    "Any notes or descriptions you added to the last output will be preserved."
)

# ──────────────────────────────────────────────────────────────────────────────
# Constants
# ──────────────────────────────────────────────────────────────────────────────

# Columns we recognise in a "previously generated" output sheet
EXISTING_OUTPUT_SHEETS = ["Filtered Logs", "Vehicle Year Counts", "Vehicle Year Fault Counts", "Recall Counts"]

# Columns in Sheet 1 that users fill in manually — must be preserved across runs
USER_COMMENT_COLS_S1 = ["TSRC Notes"]
USER_COMMENT_COLS_S2 = ["Description 1"]
USER_COMMENT_COLS_S3 = ["Description 2"]
USER_COMMENT_COLS_S4 = ["Description 3"]

# Deduplication key — uniquely identifies a complaint row
DEDUP_KEY = ["VIN", "DC PREFIX", "COMPLAINT"]

# ──────────────────────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────────────────────

def clean_str(series: pd.Series) -> pd.Series:
    return series.astype(str).str.strip().replace({"nan": pd.NA, "None": pd.NA, "": pd.NA})


def is_output_sheet(xf) -> bool:
    """Return True if the uploaded file looks like a previously generated output workbook."""
    try:
        xl = pd.ExcelFile(xf)
        return "Filtered Logs" in xl.sheet_names
    except Exception:
        return False


def read_existing_output(xf) -> dict[str, pd.DataFrame]:
    """Read all four sheets from a previously generated output workbook."""
    xl = pd.ExcelFile(xf)
    sheets = {}
    for name in EXISTING_OUTPUT_SHEETS:
        if name in xl.sheet_names:
            sheets[name] = pd.read_excel(xl, sheet_name=name, dtype=str)
    return sheets


def map_raw_to_standard(df_raw: pd.DataFrame, source_label: str) -> pd.DataFrame:
    """
    Map columns from a raw complaint xlsx (TC Logs format or Complaints format)
    into the standardised internal column names used by this app.
    """
    col_map = {
        "LOG_NO":     "TC File #",
        "VIN":        "VIN",
        "VMAKE":      "MAKE",
        "VMODEL":     "MODEL",
        "MODEL_YR":   "MODEL YEAR",
        "DC_PREFIX":  "DC PREFIX",
        "DC_NAME":    "DC NAME",
        "DC_FAULT":   "DC FAULT",
        "COMMENT":    "COMPLAINT",
        "PROVINCE":   "Province",
        "INJURY":     "Injury",
        "DATE_REP":   "Date_reported",
        "DATE_REPORT":"Date_reported",   # TC Logs uses this name
        "TSRC Notes": "TSRC Notes",
    }

    df = df_raw.rename(columns={k: v for k, v in col_map.items() if k in df_raw.columns})
    df["SOURCE"] = source_label

    # Ensure every expected column exists (fill with NA if missing)
    for col in ["TC File #", "VIN", "MAKE", "MODEL", "MODEL YEAR",
                "DC PREFIX", "DC NAME", "DC FAULT", "COMPLAINT",
                "Province", "Injury", "Date_reported", "TSRC Notes", "SOURCE"]:
        if col not in df.columns:
            df[col] = pd.NA

    return df[[
        "SOURCE", "TC File #", "VIN", "MAKE", "MODEL", "MODEL YEAR",
        "DC PREFIX", "DC NAME", "DC FAULT", "COMPLAINT",
        "Province", "Injury", "Date_reported", "TSRC Notes",
    ]]


def clean_df(df: pd.DataFrame) -> pd.DataFrame:
    for col in ["VIN", "MAKE", "MODEL", "DC PREFIX", "DC NAME", "DC FAULT", "COMPLAINT"]:
        if col in df.columns:
            df[col] = clean_str(df[col])
    return df


def style_worksheet(ws):
    header_fill  = PatternFill("solid", start_color="1F3864", end_color="1F3864")
    alt_fill     = PatternFill("solid", start_color="DCE6F1", end_color="DCE6F1")
    comment_fill = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")  # light yellow for user cols

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    body_font   = Font(name="Arial", size=10)
    thin        = Side(style="thin", color="B0B0B0")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Identify comment columns by header name
    comment_col_names = set(
        USER_COMMENT_COLS_S1 + USER_COMMENT_COLS_S2 + USER_COMMENT_COLS_S3 + USER_COMMENT_COLS_S4
    )
    comment_col_indices = {
        cell.column for cell in ws[1] if str(cell.value) in comment_col_names
    }

    for cell in ws[1]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
    ws.row_dimensions[1].height = 22

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        base_fill = alt_fill if row_idx % 2 == 0 else PatternFill()
        for cell in row:
            cell.font   = body_font
            cell.border = border
            if cell.column in comment_col_indices:
                cell.fill      = comment_fill
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            else:
                cell.fill      = base_fill
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[row_idx].height = 24

    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        header_val = str(col[0].value or "")
        max_len    = max((len(str(c.value or "")) for c in col[1:]), default=0)
        if header_val in comment_col_names or header_val == "COMPLAINT":
            ws.column_dimensions[col_letter].width = 55
        else:
            ws.column_dimensions[col_letter].width = min(max(len(header_val), max_len) + 4, 40)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


# ──────────────────────────────────────────────────────────────────────────────
# Core builder
# ──────────────────────────────────────────────────────────────────────────────

def build_workbook(existing_sheets: dict, new_dfs: list[pd.DataFrame]) -> tuple[bytes, dict]:

    # ── 1. Combine new raw files ──────────────────────────────────────────────
    df_new = pd.concat(new_dfs, ignore_index=True) if new_dfs else pd.DataFrame()

    # ── 2. Load existing Filtered Logs (carries user comments) ───────────────
    df_existing = pd.DataFrame()
    if "Filtered Logs" in existing_sheets:
        df_existing = existing_sheets["Filtered Logs"].copy()
        # Ensure user comment cols exist
        for col in USER_COMMENT_COLS_S1:
            if col not in df_existing.columns:
                df_existing[col] = pd.NA
        # Normalise column names to match internal standard
        df_existing = df_existing.rename(columns={
            "DC_PREFIX": "DC PREFIX", "DC_NAME": "DC NAME", "DC_FAULT": "DC FAULT",
        })

    # ── 3. Merge existing + new ───────────────────────────────────────────────
    if df_existing.empty and df_new.empty:
        raise ValueError("No data found — upload at least one file.")

    # Align columns
    all_cols = list(dict.fromkeys(
        ["SOURCE", "TC File #", "VIN", "MAKE", "MODEL", "MODEL YEAR",
         "DC PREFIX", "DC NAME", "DC FAULT", "COMPLAINT",
         "Province", "Injury", "Date_reported", "TSRC Notes"]
        + [c for c in (df_existing.columns.tolist() if not df_existing.empty else [])]
    ))

    def reindex(df):
        for c in all_cols:
            if c not in df.columns:
                df[c] = pd.NA
        return df[all_cols]

    df_existing = reindex(df_existing) if not df_existing.empty else pd.DataFrame(columns=all_cols)
    df_new      = reindex(df_new)      if not df_new.empty      else pd.DataFrame(columns=all_cols)

    df_combined = pd.concat([df_existing, df_new], ignore_index=True)
    df_combined = clean_df(df_combined)

    # ── 4. Deduplicate, keeping the FIRST occurrence (existing rows win so
    #       user comments are preserved) ──────────────────────────────────────
    dedup_cols = [c for c in DEDUP_KEY if c in df_combined.columns]
    df_combined = df_combined.drop_duplicates(subset=dedup_cols, keep="first").reset_index(drop=True)

    # ── 5. Restore user comments from existing sheet for any row that survived
    #       (they're already there since existing rows came first) ─────────────
    #    For rows that came from df_new but match an existing row's dedup key,
    #    the existing row already took priority above — nothing extra needed.

    # ── 6. Sheet 1: Filtered Logs ─────────────────────────────────────────────
    s1_cols = [
        "SOURCE", "TC File #", "VIN", "MAKE", "MODEL", "MODEL YEAR",
        "DC PREFIX", "DC NAME", "DC FAULT", "COMPLAINT",
        "Province", "Injury", "Date_reported",
        "TSRC Notes",
    ]
    df_s1 = df_combined[[c for c in s1_cols if c in df_combined.columns]].copy()

    # Sort: most-complained model first, then alpha
    model_counts = df_s1["MODEL"].value_counts(dropna=False)
    df_s1["_mc"] = df_s1["MODEL"].map(model_counts)
    df_s1 = df_s1.sort_values(
        by=["_mc", "MODEL", "MODEL YEAR", "DC PREFIX", "DC FAULT"],
        ascending=[False, True, True, True, True]
    ).drop(columns=["_mc"]).reset_index(drop=True)

    # ── 7. Sheet 2: Vehicle Year Counts ──────────────────────────────────────
    df_s2 = (
        df_combined.dropna(subset=["VIN", "MAKE", "MODEL", "MODEL YEAR"])
        .drop_duplicates(subset=["VIN", "MAKE", "MODEL", "MODEL YEAR"])
        .groupby(["MAKE", "MODEL", "MODEL YEAR"])
        .size()
        .reset_index(name="TOTAL")
        .sort_values(["TOTAL", "MAKE", "MODEL", "MODEL YEAR"], ascending=[False, True, True, True])
        .reset_index(drop=True)
    )
    df_s2["Description 1"] = pd.NA
    # Restore existing user descriptions
    if "Vehicle Year Counts" in existing_sheets:
        df_ex2 = existing_sheets["Vehicle Year Counts"].copy()
        if "Description 1" in df_ex2.columns:
            df_ex2["MODEL YEAR"] = df_ex2["MODEL YEAR"].astype(str).str.strip()
            df_s2["MODEL YEAR"]  = df_s2["MODEL YEAR"].astype(str).str.strip()
            merge_key2 = ["MAKE", "MODEL", "MODEL YEAR"]
            df_s2 = df_s2.merge(
                df_ex2[merge_key2 + ["Description 1"]].rename(columns={"Description 1": "_desc1"}),
                on=merge_key2, how="left"
            )
            mask = df_s2["_desc1"].notna()
            df_s2.loc[mask, "Description 1"] = df_s2.loc[mask, "_desc1"]
            df_s2 = df_s2.drop(columns=["_desc1"])

    # ── 8. Sheet 3: Vehicle Year Fault Counts ────────────────────────────────
    df_s3 = (
        df_combined.dropna(subset=["MAKE", "MODEL", "MODEL YEAR", "DC FAULT"])
        .groupby(["MAKE", "MODEL", "MODEL YEAR", "DC FAULT"])
        .size()
        .reset_index(name="TOTAL")
        .sort_values(["TOTAL", "MAKE", "MODEL", "MODEL YEAR", "DC FAULT"], ascending=[False, True, True, True, True])
        .reset_index(drop=True)
    )
    df_s3["Description 2"] = pd.NA
    if "Vehicle Year Fault Counts" in existing_sheets:
        df_ex3 = existing_sheets["Vehicle Year Fault Counts"].copy()
        if "Description 2" in df_ex3.columns:
            df_ex3["MODEL YEAR"] = df_ex3["MODEL YEAR"].astype(str).str.strip()
            df_s3["MODEL YEAR"]  = df_s3["MODEL YEAR"].astype(str).str.strip()
            merge_key3 = ["MAKE", "MODEL", "MODEL YEAR", "DC FAULT"]
            df_s3 = df_s3.merge(
                df_ex3[merge_key3 + ["Description 2"]].rename(columns={"Description 2": "_desc2"}),
                on=merge_key3, how="left"
            )
            mask = df_s3["_desc2"].notna()
            df_s3.loc[mask, "Description 2"] = df_s3.loc[mask, "_desc2"]
            df_s3 = df_s3.drop(columns=["_desc2"])

    # ── 9. Sheet 4: Recall Counts ─────────────────────────────────────────────
    recall_pattern = re.compile(r"\b\d{2}[A-Za-z]\b")
    recall_matches = []
    for comment in df_combined["COMPLAINT"].dropna().astype(str):
        for m in recall_pattern.findall(comment.upper()):
            recall_matches.append(m)

    if recall_matches:
        df_s4 = (
            pd.DataFrame({"RECALL": recall_matches})
            .assign(RECALL=lambda d: "Recall " + d["RECALL"])
            .groupby("RECALL").size().reset_index(name="TOTAL")
            .sort_values(["TOTAL", "RECALL"], ascending=[False, True])
            .reset_index(drop=True)
        )
    else:
        df_s4 = pd.DataFrame(columns=["RECALL", "TOTAL"])

    df_s4["Description 3"] = pd.NA
    if "Recall Counts" in existing_sheets:
        df_ex4 = existing_sheets["Recall Counts"].copy()
        if "Description 3" in df_ex4.columns:
            df_s4 = df_s4.merge(
                df_ex4[["RECALL", "Description 3"]].rename(columns={"Description 3": "_desc3"}),
                on="RECALL", how="left"
            )
            mask = df_s4["_desc3"].notna()
            df_s4.loc[mask, "Description 3"] = df_s4.loc[mask, "_desc3"]
            df_s4 = df_s4.drop(columns=["_desc3"])

    # ── 10. Write workbook ────────────────────────────────────────────────────
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_s1.to_excel(writer, sheet_name="Filtered Logs",             index=False)
        df_s2.to_excel(writer, sheet_name="Vehicle Year Counts",       index=False)
        df_s3.to_excel(writer, sheet_name="Vehicle Year Fault Counts", index=False)
        df_s4.to_excel(writer, sheet_name="Recall Counts",             index=False)

    buf.seek(0)
    wb = load_workbook(buf)
    for name in wb.sheetnames:
        style_worksheet(wb[name])

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    return out.getvalue(), {
        "total_rows": len(df_s1),
        "vehicle_year_rows": len(df_s2),
        "fault_rows": len(df_s3),
        "recall_rows": len(df_s4),
    }


# ──────────────────────────────────────────────────────────────────────────────
# UI
# ──────────────────────────────────────────────────────────────────────────────

st.markdown("### Step 1 — Previous output sheet *(optional on first run)*")
existing_upload = st.file_uploader(
    "Upload previously generated output workbook (TC_Logs_Filtered.xlsx)",
    type=["xlsx"],
    accept_multiple_files=False,
    key="existing"
)

st.markdown("### Step 2 — New complaint files")
new_uploads = st.file_uploader(
    "Upload one or more new raw complaint Excel files",
    type=["xlsx"],
    accept_multiple_files=True,
    key="new_files"
)

if st.button("Generate workbook", type="primary"):
    if not existing_upload and not new_uploads:
        st.error("Please upload at least one file.")
    else:
        try:
            with st.spinner("Processing..."):

                # Load existing output (if provided)
                existing_sheets = {}
                if existing_upload:
                    existing_upload.seek(0)
                    if is_output_sheet(existing_upload):
                        existing_upload.seek(0)
                        existing_sheets = read_existing_output(existing_upload)
                    else:
                        st.warning(
                            "The file uploaded under 'Previous output' doesn't look like a "
                            "generated output sheet — treating it as a raw complaint file."
                        )
                        existing_upload.seek(0)
                        new_uploads = list(new_uploads) + [existing_upload]

                # Process each new raw complaint file
                new_dfs = []
                for f in (new_uploads or []):
                    f.seek(0)
                    df_raw = pd.read_excel(f, dtype=str)
                    label  = f.name.replace(".xlsx", "")
                    new_dfs.append(map_raw_to_standard(df_raw, label))

                file_bytes, stats = build_workbook(existing_sheets, new_dfs)

            st.success("Workbook created successfully.")

            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Total complaint rows",  stats["total_rows"])
            c2.metric("Vehicle/Year combos",   stats["vehicle_year_rows"])
            c3.metric("Vehicle/Year/Fault",    stats["fault_rows"])
            c4.metric("Recall codes found",    stats["recall_rows"])

            st.info(
                "💡 Yellow-highlighted columns (TSRC Notes, Description 1/2/3) are for your notes. "
                "Fill them in the downloaded file — they'll be preserved next time you upload it."
            )

            st.download_button(
                label="⬇️ Download TC_Logs_Filtered.xlsx",
                data=file_bytes,
                file_name="TC_Logs_Filtered.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        except Exception as e:
            st.error(f"Error: {e}")
            st.exception(e)
